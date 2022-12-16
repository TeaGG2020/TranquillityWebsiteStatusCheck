#Imports /// Импорты
import os
import requests
import openpyxl
import warnings
import pandas as pd
from itertools import count
from openpyxl import load_workbook
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter

#Ignore errors related to SSL certificates /// Игнорируем ошибки связанные с SSL сертификатами
warnings.filterwarnings("ignore")

#Variables /// Переменные
countAll = 0
countGood = 0
countBad = 0

#Enter file name /// Ввод названия файла
print("Enter the name of the file to be checked! (For example: Input.csv)")
file_name = str(input())


#Convert from csv to xlsx for convenience /// Конвертация из Csv в Xlsx для удобства
dir_path = os.path.dirname(os.path.realpath(__file__))
cvsDataframe = pd.read_csv(dir_path+'\\'+file_name)
resultExcelFile = pd.ExcelWriter(dir_path+'\\Output.xlsx')
cvsDataframe.to_excel(resultExcelFile, index=False)
resultExcelFile.save()
wb = openpyxl.load_workbook(dir_path+'\\Output.xlsx')
sh = wb.active

#Connect function, with the ability to reconnect /// Функция коннекта, с возможность реконнекта
session = requests.Session()
retry = Retry(connect=5,backoff_factor=0.5)
adapter = HTTPAdapter(max_retries=retry)
session.mount('http://', adapter)
session.mount('https://', adapter)

#Main Checking /// Основная проверка
def checking(value_cell_name):
    global countAll
    global countGood
    global countBad
    for i in count(start=2):
        try:
            value = session.get('http://'+sh[value_cell_name+str(i)].value,verify=False)
        except:
            value = 404
        
        #Intermediate results /// Промежуточные результаты
        def prints(type):
            print('Checking website'+type+' in '+value_cell_name+str(i)+' - ',end='') 
        if value_cell_name == "C":
            prints('')
        elif value_cell_name == "F":
            prints(' price')
        elif value_cell_name == "H":
            prints(' trial')
               
        if value == 404 or value.status_code == 404:
            workbook = load_workbook(filename=dir_path+'\\Output.xlsx')
            sheet = workbook.active
            if value_cell_name == "C":
                sheet[value_cell_name+str(i)] = 404
            elif value_cell_name == "F":
                sheet['G'+str(i)] = False
            elif value_cell_name == "H":
                sheet['I'+str(i)] = False
            workbook.save(filename=dir_path+'\\Output.xlsx')
            print("Bad")
            countAll = countAll + 1
            countBad = countBad + 1
        elif value.status_code == 200 or value.status_code == 403:
            if value_cell_name == "F" or value_cell_name == "H":
                workbook = load_workbook(filename=dir_path+'\\Output.xlsx')
                sheet = workbook.active
                if value_cell_name == "F":
                    sheet['G'+str(i)] = True
                elif value_cell_name == "H":
                    sheet['I'+str(i)] = True
                workbook.save(filename=dir_path+'\\Output.xlsx')
            print("Good")
            countAll = countAll + 1
            countGood = countGood + 1
            continue
        elif value.status_code != 200 or value.status_code != 404 or value.status_code != 403:
            print("\nResults:","\nAll - ",countAll,"\nGood - ",countGood,"\nBad - ",countBad)
            break

#Website home page check /// Проверка главной страницы сайта
checking("C")

#Website price page check /// Проверка страницы с ценой на сайте
checking("F")

#Website trial page check /// Проверка пробной страницы сайта
checking("H")

#Let's create a csv file /// Создадим Csv файл
data_xls = pd.read_excel(dir_path+'\\Output.xlsx', 'Sheet1', dtype=str, index_col=None)
data_xls.to_csv(dir_path+'\\Output.csv', encoding='utf-8', index=False)